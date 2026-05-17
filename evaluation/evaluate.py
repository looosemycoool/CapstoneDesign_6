import os
import re
import json
import time
import traceback
import subprocess
import importlib.util
from datetime import datetime

import pandas as pd
import openpyxl


def _get_git_hash() -> str:
    """현재 commit 의 짧은 hash (재현성 위해 xlsx 메타데이터에 기록)."""
    try:
        h = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL, timeout=2,
        ).decode().strip()
        return h or "unknown"
    except Exception:
        return "unknown"


def _get_eval_metadata() -> dict:
    """평가 재현성을 위한 모델/하이퍼파라미터/git hash 메타데이터.
    값이 바뀌면 여기 직접 갱신 (런타임 동적 import 보다 단순/명시적이 낫다)."""
    return {
        "Generation 모델":  "solar-pro3",
        "Generation temp":  "0",
        "Judge 모델":       f"{os.getenv('EVAL_JUDGE_MODEL', 'solar-pro3')} (temp=0)",
        "Embedding 모델":   "embedding-passage",
        "chunk_size":       "500 (실측)",
        "max_relations":    "3",
        "n_results":        "5",
        "git hash":         _get_git_hash(),
    }
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 깔끔한 xlsx 저장 ──────────────────────────────────────
def _save_clean_xlsx(xlsx_path, rows, collection_name, experiment_id, elapsed, timestamp):
    """
    시트 구성:
      ① 요약  : 모델 정보 + Excel 수식 자동 집계
      ② 결과  : 핵심 11열, 성공/실패 색상, 헤더 freeze
    """
    # ── 색상 / 폰트 ──────────────────────────────────────
    BLUE_FILL    = PatternFill("solid", fgColor="1F497D")
    TEAL_FILL    = PatternFill("solid", fgColor="17375E")
    SUCCESS_FILL = PatternFill("solid", fgColor="E2EFDA")
    FAIL_FILL    = PatternFill("solid", fgColor="FFDDC1")
    Q_FILL       = PatternFill("solid", fgColor="F5F5F5")
    H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    B_FONT  = Font(bold=True, name="Arial", size=9)
    N_FONT  = Font(name="Arial", size=9)
    AL_C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_TL   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    AL_TC   = Alignment(horizontal="center", vertical="top")

    n = len(rows)  # 질문 수

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── ① 요약 시트 ──────────────────────────────────────
    ws1 = wb.create_sheet("① 요약")

    def label_row(ws, row, label, value, label_fill=None):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = B_FONT
        vc.font = N_FONT
        lc.alignment = AL_C
        vc.alignment = AL_TL
        if label_fill:
            lc.fill = label_fill
            lc.font = H_FONT

    # 모델 정보 블록
    ws1.merge_cells("A1:B1")
    title = ws1["A1"]
    title.value = "📊 평가 결과 요약"
    title.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    title.fill = TEAL_FILL
    title.alignment = AL_C
    ws1.row_dimensions[1].height = 28

    info = [
        ("실험 ID",       experiment_id),
        ("벡터 컬렉션",   collection_name),
        ("총 질문 수",    n),
        ("평가 일시",     timestamp),
        ("소요 시간(초)", elapsed),
    ]
    # 재현성용 메타데이터 (모델/하이퍼파라미터/git hash) 추가
    info.extend(_get_eval_metadata().items())

    for i, (label, val) in enumerate(info, start=2):
        label_row(ws1, i, label, val)

    # 동적 행 계산: title=1, info=2..(1+len), 빈 줄, 성능 지표 헤더, metrics
    metrics_header_row = 2 + len(info) + 1  # info 끝 다음 행
    metrics_start_row = metrics_header_row + 1

    ws1.merge_cells(start_row=metrics_header_row, start_column=1,
                    end_row=metrics_header_row, end_column=2)
    ws1.cell(row=metrics_header_row, column=1, value="📈 성능 지표")
    ws1.cell(row=metrics_header_row, column=1).font = H_FONT
    ws1.cell(row=metrics_header_row, column=1).fill = BLUE_FILL
    ws1.cell(row=metrics_header_row, column=1).alignment = AL_C
    ws1.row_dimensions[metrics_header_row].height = 22

    # 새 컬럼 인덱스 (1-based): F=Vector 정답, J=Hybrid 정답, M=Graph 수
    metrics = [
        ("Vector 정답 수",    f"=COUNTIF('② 결과'!F2:F{n+1},\"정답\")"),
        ("Vector 정답률",     f"=COUNTIF('② 결과'!F2:F{n+1},\"정답\")/COUNTA('② 결과'!F2:F{n+1})"),
        ("Hybrid 정답 수",    f"=COUNTIF('② 결과'!J2:J{n+1},\"정답\")"),
        ("Hybrid 정답률",     f"=COUNTIF('② 결과'!J2:J{n+1},\"정답\")/COUNTA('② 결과'!J2:J{n+1})"),
        ("평균 Graph 관계 수", f"=AVERAGE('② 결과'!M2:M{n+1})"),
    ]
    for i, (label, formula) in enumerate(metrics, start=metrics_start_row):
        lc = ws1.cell(row=i, column=1, value=label)
        vc = ws1.cell(row=i, column=2, value=formula)
        lc.font = B_FONT
        lc.alignment = AL_C
        vc.font = N_FONT
        vc.alignment = AL_TC
        if "정답률" in label:
            vc.number_format = "0.0%"
        elif "평균" in label:
            vc.number_format = "0.00"

    # 테두리
    thin = Side(style="thin", color="CCCCCC")
    last_row = metrics_start_row + len(metrics)
    for row in ws1.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 30

    # ── ② 결과 시트 ──────────────────────────────────────
    ws2 = wb.create_sheet("② 결과")

    headers = [
        "No", "번호", "분류", "질문", "정답",
        "Vector 정답", "Vector 답변", "Vector 사유", "Vector 오류",
        "Hybrid 정답", "Hybrid 답변", "Hybrid 사유", "Graph 수", "Hybrid 오류",
    ]
    col_widths = [5, 6, 10, 42, 32, 10, 42, 30, 20, 10, 42, 30, 8, 20]
    SUCCESS_COLS = {6, 10}  # Vector 정답, Hybrid 정답 (1-based)
    GRAPH_COL = 13           # Graph 수 컬럼 인덱스

    # 헤더
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font      = H_FONT
        cell.fill      = BLUE_FILL
        cell.alignment = AL_C
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[1].height = 30

    # 데이터
    for idx, r in enumerate(rows, start=2):
        # vector_success/hybrid_success 는 이제 "정답 여부" (judge LLM 결과 또는 fallback)
        v_str = "정답" if r.get("vector_success") is True else "오답"
        h_str = "정답" if r.get("hybrid_success") is True else "오답"

        data = [
            r.get("no", ""),
            r.get("id", ""),
            r.get("category", ""),
            r.get("question", ""),
            r.get("ground_truth", ""),
            v_str,
            r.get("vector_answer", ""),
            r.get("vector_judge_reason", ""),
            r.get("vector_error", ""),
            h_str,
            r.get("hybrid_answer", ""),
            r.get("hybrid_judge_reason", ""),
            r.get("hybrid_graph_count", 0),
            r.get("hybrid_error", ""),
        ]

        for col_idx, val in enumerate(data, start=1):
            cell = ws2.cell(row=idx, column=col_idx, value=val)
            cell.font = N_FONT

            if col_idx in SUCCESS_COLS:
                cell.fill      = SUCCESS_FILL if val == "정답" else FAIL_FILL
                cell.alignment = AL_TC
            elif col_idx == 4:  # 질문 열 배경
                cell.fill      = Q_FILL
                cell.alignment = AL_TL
            elif col_idx == GRAPH_COL:  # Graph 수
                cell.alignment = AL_TC
            else:
                cell.alignment = AL_TL

        ws2.row_dimensions[idx].height = 15

    ws2.freeze_panes = "F2"  # No·번호·분류·질문까지 고정, 성공/실패부터 스크롤

    # ── ③ 세부지표 시트 (논문 분석용 raw fields + 계산 metric) ──────────
    _add_detail_sheet(wb, rows, H_FONT, B_FONT, N_FONT, AL_C, AL_TL, AL_TC,
                      BLUE_FILL, SUCCESS_FILL, FAIL_FILL, Q_FILL)

    wb.save(xlsx_path)


def _xlsx_safe(val):
    """openpyxl 이 거부하는 control character 제거.
    raw context/answer 에 LLM 이 넣은 illegal char (예: \\x00) 를 sanitize."""
    if not isinstance(val, str):
        return val
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    return ILLEGAL_CHARACTERS_RE.sub("", val)


def _add_detail_sheet(wb, rows, H_FONT, B_FONT, N_FONT, AL_C, AL_TL, AL_TC,
                      BLUE_FILL, SUCCESS_FILL, FAIL_FILL, Q_FILL):
    """논문 분석용 raw fields 가 모두 들어있는 세부지표 시트.
    qa_dataset 에서 persona 도 join 해서 페르소나 단위 분석 가능.
    """
    # qa_dataset 으로 persona 매핑 (id 기준)
    persona_by_id = {}
    try:
        with open(QA_DATASET_PATH, encoding="utf-8") as f:
            for q in json.load(f):
                qid = str(q.get("id", "")).strip()
                if qid:
                    persona_by_id[qid] = q.get("persona", "")
    except Exception:
        pass

    ws = wb.create_sheet("③ 세부지표")

    # 컬럼 헤더 + 너비 정의 (paper raw analysis 용)
    columns = [
        ("No", 5),
        ("번호", 6),
        ("분류", 12),
        ("페르소나", 18),
        ("질문", 40),
        ("정답", 40),
        ("정답 길이", 10),
        # ─ Vector ─
        ("V_verdict", 10),
        ("V_정답여부", 10),
        ("V_답변", 40),
        ("V_답변길이", 10),
        ("V_judge사유", 30),
        ("V_sources", 30),
        ("V_sources_count", 10),
        ("V_scored_sources", 30),
        ("V_context_preview", 40),
        ("V_error", 20),
        # ─ Hybrid ─
        ("H_verdict", 10),
        ("H_정답여부", 10),
        ("H_답변", 40),
        ("H_답변길이", 10),
        ("H_judge사유", 30),
        ("H_sources", 30),
        ("H_sources_count", 10),
        ("H_scored_sources", 30),
        ("H_context_preview", 40),
        ("H_graph_count", 10),
        ("H_graph_preview", 30),
        ("H_error", 20),
    ]
    headers = [c[0] for c in columns]
    widths = [c[1] for c in columns]
    SUCCESS_COLS_DETAIL = {9, 19}  # V_정답여부, H_정답여부 (1-based)

    from openpyxl.utils import get_column_letter
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = H_FONT
        cell.fill = BLUE_FILL
        cell.alignment = AL_C
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    for idx, r in enumerate(rows, start=2):
        v_str = "정답" if r.get("vector_success") is True else "오답"
        h_str = "정답" if r.get("hybrid_success") is True else "오답"
        v_sources_str = r.get("vector_sources", "") or ""
        h_sources_str = r.get("hybrid_sources", "") or ""
        v_src_count = len([s for s in v_sources_str.split(" | ") if s.strip()])
        h_src_count = len([s for s in h_sources_str.split(" | ") if s.strip()])

        data = [
            r.get("no", ""),
            r.get("id", ""),
            r.get("category", ""),
            persona_by_id.get(str(r.get("id", "")), ""),
            r.get("question", ""),
            r.get("ground_truth", ""),
            len(r.get("ground_truth", "") or ""),
            # Vector
            r.get("vector_judge_verdict", ""),
            v_str,
            r.get("vector_answer", ""),
            len(r.get("vector_answer", "") or ""),
            r.get("vector_judge_reason", ""),
            v_sources_str,
            v_src_count,
            r.get("vector_scored_sources", ""),
            r.get("vector_context_preview", ""),
            r.get("vector_error", ""),
            # Hybrid — columns 정의 순서와 동일하게 유지 (CodeRabbit 지적 fix):
            # H_verdict, H_정답여부, H_답변, H_답변길이, H_judge사유, H_sources,
            # H_sources_count, H_scored_sources, H_context_preview,
            # H_graph_count, H_graph_preview, H_error
            r.get("hybrid_judge_verdict", ""),
            h_str,
            r.get("hybrid_answer", ""),
            len(r.get("hybrid_answer", "") or ""),
            r.get("hybrid_judge_reason", ""),
            h_sources_str,
            h_src_count,
            r.get("hybrid_scored_sources", ""),
            r.get("hybrid_context_preview", ""),
            r.get("hybrid_graph_count", 0),
            r.get("hybrid_graph_preview", ""),
            r.get("hybrid_error", ""),
        ]
        for col_idx, val in enumerate(data, start=1):
            safe_val = _xlsx_safe(val)
            cell = ws.cell(row=idx, column=col_idx, value=safe_val)
            cell.font = N_FONT
            if col_idx in SUCCESS_COLS_DETAIL:
                cell.fill = SUCCESS_FILL if safe_val == "정답" else FAIL_FILL
                cell.alignment = AL_TC
            elif col_idx == 5:  # 질문 배경
                cell.fill = Q_FILL
                cell.alignment = AL_TL
            else:
                cell.alignment = AL_TL
        ws.row_dimensions[idx].height = 15

    ws.freeze_panes = "G2"  # No·번호·분류·페르소나·질문까지 고정


# ── 경로 설정 ─────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EVAL_DIR = os.path.join(BASE_DIR, "evaluation")
RESULTS_DIR = os.path.join(EVAL_DIR, "results")
QA_DATASET_PATH = os.path.join(EVAL_DIR, "qa_dataset.json")
HYBRID_RAG_PATH = os.path.join(BASE_DIR, "pipeline", "04_hybrid_rag.py")

os.makedirs(RESULTS_DIR, exist_ok=True)


# ── 하이브리드 모듈 동적 로드 ─────────────────────────────
def load_hybrid_module():
    spec = importlib.util.spec_from_file_location("hybrid_rag_module", HYBRID_RAG_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


# ── QA 데이터셋 로드 ──────────────────────────────────────
def load_qa_dataset(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"QA 데이터셋이 없습니다: {path}")

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("qa_dataset.json은 list 형태여야 합니다.")

    return data


# ── QA 데이터셋의 다양한 키 형태 지원 ─────────────────────
def extract_question(item):
    for key in ["question", "query", "질문"]:
        value = item.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def extract_ground_truth(item):
    for key in ["ground_truth", "answer", "reference_answer", "expected_answer", "정답"]:
        value = item.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def extract_id(item, index):
    for key in ["id", "qid", "question_id", "번호"]:
        value = item.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return str(index + 1)


def extract_category(item):
    for key in ["category", "type", "persona", "source", "분류"]:
        value = item.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


# ── 결과 요약용 유틸 ──────────────────────────────────────
def safe_join(items, sep=" | "):
    return sep.join([str(x) for x in items if x is not None and str(x).strip()])


def summarize_vector_docs(vector_docs):
    sources = []
    scored_sources = []

    for doc in vector_docs or []:
        source = doc.get("source", "")
        score = doc.get("score", "")
        if source:
            sources.append(source)
            scored_sources.append(f"{source} ({score})")

    return {
        "sources": safe_join(dict.fromkeys(sources).keys()),
        "scored_sources": safe_join(scored_sources),
    }


def summarize_graph_relations(graph_relations, max_preview=10):
    preview = []
    for rel in (graph_relations or [])[:max_preview]:
        from_node = rel.get("from", "")
        relation = rel.get("relation", "")
        to_node = rel.get("to", "")
        if from_node and relation and to_node:
            preview.append(f"{from_node} --[{relation}]--> {to_node}")

    return {
        "count": len(graph_relations or []),
        "preview": "\n".join(preview)
    }


def truncate_text(text, limit=1500):
    if not text:
        return ""
    text = str(text)
    if len(text) <= limit:
        return text
    return text[:limit] + "...[truncated]"


# ── LLM-as-judge: 답변 정답 여부 채점 ─────────────────────
JUDGE_MODEL = "solar-pro3"  # generation 모델과 동급 이상이어야 판정 신뢰도 확보
JUDGE_TEMPERATURE = 0
JUDGE_BASE_URL = "https://api.upstage.ai/v1"


def get_judge_llm():
    """채점용 클라이언트 (Upstage solar-pro2, OpenAI SDK 호환 endpoint).
    환경변수 EVAL_JUDGE_MODEL 로 모델 변경 가능."""
    from openai import OpenAI
    api_key = os.getenv("UPSTAGE_API_KEY")
    if not api_key:
        raise RuntimeError("UPSTAGE_API_KEY 가 .env 에 없음 — judge 사용 불가")
    return OpenAI(api_key=api_key, base_url=JUDGE_BASE_URL)


def judge_answer(question, ground_truth, predicted, judge_llm):
    """LLM 으로 답변이 ground_truth 와 의미상 일치하는지 판단.
    반환: (verdict, reason). verdict ∈ {"correct", "incorrect", "skipped", "error"}.
    """
    if not predicted or not str(predicted).strip():
        return "incorrect", "답변 없음"
    if not ground_truth or not str(ground_truth).strip():
        return "skipped", "정답(ground_truth) 미제공"

    prompt = (
        "아래 RAG 챗봇 답변이 정답과 의미상 일치하는지 판단하세요.\n\n"
        "## 판정 기준\n"
        "correct  : 정답의 핵심 사실이 답변에 모두 포함되어 있고 정확함.\n"
        "           정답보다 더 상세하거나 추가 정보가 있어도 핵심이 맞으면 correct.\n"
        "incorrect: 정답의 핵심 사실 중 하나라도 (1) 틀린 값으로 기재되거나 (2) 완전히 누락됨.\n"
        "           또는 질문 대상과 다른 항목·프로그램의 값을 핵심 답변으로 제시한 경우.\n\n"
        "## 수치 판정 (엄격 적용)\n"
        "- 점수·학점·금액·기간 등 수치는 정답과 정확히 일치해야 correct.\n"
        "- 그럴듯하지만 다른 수치(예: 정답 83점 → 답변 94점, 정답 2억 → 답변 3억)는 incorrect.\n"
        "- 답변이 아무리 길고 논리적으로 보여도 핵심 수치가 하나라도 틀리면 incorrect.\n"
        "- 정답과 다른 프로그램·제도·트랙에서 가져온 수치를 핵심 답변으로 제시한 경우 incorrect.\n\n"
        "## verbosity 편향 방지 (필수)\n"
        "- 답변 길이·설명 분량·근거 나열 방식은 판정에 절대 영향 없음.\n"
        "- 짧고 직접적인 답변과 길고 상세한 답변을 반드시 동일 기준으로 판정.\n"
        "- 답변이 길고 자세해도 핵심 수치나 사실이 틀리면 incorrect.\n"
        "- 답변이 짧아도 핵심 사실이 맞으면 correct.\n\n"
        "## 추가 정보 처리\n"
        "- 추가 정보(정답에 없지만 사실에 부합하는 내용)만으로는 incorrect 판정 금지.\n"
        "- 사소한 표현·어순 차이는 correct.\n\n"
        f"질문: {question}\n"
        f"정답: {ground_truth}\n"
        f"답변: {predicted}\n\n"
        "판정 순서: ① 정답의 핵심 수치·사실을 먼저 확인 → ② 답변에서 해당 수치·사실 대조 "
        "→ ③ 하나라도 틀리거나 누락되면 incorrect, 모두 맞으면 correct.\n\n"
        "다음 JSON 한 줄로만 답하세요:\n"
        '{"verdict": "correct" 또는 "incorrect", "reason": "한 문장"}'
    )
    try:
        model_name = os.getenv("EVAL_JUDGE_MODEL", JUDGE_MODEL)
        response = judge_llm.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": prompt}],
            temperature=JUDGE_TEMPERATURE,
        )
        raw = response.choices[0].message.content.strip()
        m = re.search(r"\{[^{}]*\}", raw, re.DOTALL)
        if m:
            try:
                obj = json.loads(m.group(0))
                v = str(obj.get("verdict", "")).strip().lower()
                rsn = str(obj.get("reason", "")).strip()
                if v in ("correct", "incorrect"):
                    return v, rsn
            except json.JSONDecodeError:
                pass
        low = raw.lower()
        if "incorrect" in low:
            return "incorrect", raw[:200]
        if "correct" in low:
            return "correct", raw[:200]
        return "error", f"파싱 실패: {raw[:200]}"
    except Exception as e:
        return "error", f"{type(e).__name__}: {e}"


# ── 질문 1개 평가 ────────────────────────────────────────
def evaluate_one(module, item, index, judge_llm=None):
    qid = extract_id(item, index)
    category = extract_category(item)
    question = extract_question(item)
    ground_truth = extract_ground_truth(item)

    row = {
        "id": qid,
        "category": category,
        "question": question,
        "ground_truth": ground_truth,

        # vector_success: True 면 정답으로 판정 (judge_llm 미설정 시 실행 성공만 의미)
        "vector_success": False,
        "vector_judge_verdict": "skipped",
        "vector_judge_reason": "",
        "vector_answer": "",
        "vector_sources": "",
        "vector_scored_sources": "",
        "vector_context_preview": "",
        "vector_error": "",

        "hybrid_success": False,
        "hybrid_judge_verdict": "skipped",
        "hybrid_judge_reason": "",
        "hybrid_answer": "",
        "hybrid_sources": "",
        "hybrid_scored_sources": "",
        "hybrid_graph_count": 0,
        "hybrid_graph_preview": "",
        "hybrid_context_preview": "",
        "hybrid_error": "",
    }

    if not question:
        row["vector_error"] = "질문 없음"
        row["hybrid_error"] = "질문 없음"
        return row

    # Vector Only — 실행
    vector_answer = ""
    try:
        vector_result = module.vector_only_rag(question, verbose=False)
        vector_docs = vector_result.get("vector_docs", [])
        vector_summary = summarize_vector_docs(vector_docs)

        vector_answer = vector_result.get("answer", "")
        row["vector_answer"] = vector_answer
        row["vector_sources"] = vector_summary["sources"]
        row["vector_scored_sources"] = vector_summary["scored_sources"]
        row["vector_context_preview"] = truncate_text(vector_result.get("context", ""))
    except Exception as e:
        row["vector_error"] = f"{type(e).__name__}: {e}"

    # Vector — 채점
    if row["vector_error"]:
        row["vector_judge_verdict"] = "error"
        row["vector_judge_reason"] = "파이프라인 오류"
    elif judge_llm is not None:
        verdict, reason = judge_answer(question, ground_truth, vector_answer, judge_llm)
        row["vector_judge_verdict"] = verdict
        row["vector_judge_reason"] = reason
        row["vector_success"] = (verdict == "correct")
    else:
        # judge LLM 미설정: 이전 동작 유지 (실행 성공 = success)
        row["vector_success"] = True
        row["vector_judge_reason"] = "judge 미설정"

    # Hybrid — 실행
    hybrid_answer = ""
    try:
        hybrid_result = module.hybrid_rag(question, verbose=False)
        hybrid_docs = hybrid_result.get("vector_docs", [])
        hybrid_summary = summarize_vector_docs(hybrid_docs)
        graph_summary = summarize_graph_relations(hybrid_result.get("graph_relations", []))

        hybrid_answer = hybrid_result.get("answer", "")
        row["hybrid_answer"] = hybrid_answer
        row["hybrid_sources"] = hybrid_summary["sources"]
        row["hybrid_scored_sources"] = hybrid_summary["scored_sources"]
        row["hybrid_graph_count"] = graph_summary["count"]
        row["hybrid_graph_preview"] = graph_summary["preview"]
        row["hybrid_context_preview"] = truncate_text(hybrid_result.get("context", ""))
    except Exception as e:
        row["hybrid_error"] = f"{type(e).__name__}: {e}"

    # Hybrid — 채점
    if row["hybrid_error"]:
        row["hybrid_judge_verdict"] = "error"
        row["hybrid_judge_reason"] = "파이프라인 오류"
    elif judge_llm is not None:
        verdict, reason = judge_answer(question, ground_truth, hybrid_answer, judge_llm)
        row["hybrid_judge_verdict"] = verdict
        row["hybrid_judge_reason"] = reason
        row["hybrid_success"] = (verdict == "correct")
    else:
        row["hybrid_success"] = True
        row["hybrid_judge_reason"] = "judge 미설정"

    return row


# ── 평가 범위 입력 ────────────────────────────────────────
def _prompt_eval_range(total: int):
    """사용자로부터 평가 범위를 입력받아 (start_idx, end_idx) 0-based 슬라이스 인덱스 반환."""
    print()
    print("=" * 58)
    print(f"  평가 범위 선택  (QA 데이터셋 총 {total}개 항목)")
    print("=" * 58)
    print("  입력 형식:")
    print(f"    0      전체 평가 (1~{total}번, {total}개 모두)")
    print(f"    N      N번 항목 1개만        예) 5")
    print(f"    N,M    N번~M번 범위 평가     예) 1,50  →  1~50번")
    print(f"    N,     N번부터 끝까지        예) 51,   →  51~{total}번")
    print(f"    ,M     처음부터 M번까지      예) ,30   →  1~30번")
    print("=" * 58)

    while True:
        raw = input("  입력 > ").strip()
        if not raw:
            print("  [오류] 값을 입력하세요.")
            continue

        try:
            if raw == "0":
                print(f"  → 전체 {total}개 항목을 평가합니다.\n")
                return 0, total

            if "," in raw:
                left, right = raw.split(",", 1)
                left, right = left.strip(), right.strip()
                start = (int(left) - 1) if left else 0
                end   = int(right) if right else total
            else:
                n = int(raw)
                start, end = n - 1, n

            if start < 0 or end > total or start >= end:
                print(f"  [오류] 1~{total} 범위 내에서 start < end 가 되도록 지정하세요.")
                continue

            count = end - start
            print(f"  → {start + 1}번~{end}번 항목을 평가합니다. ({count}개)\n")
            return start, end

        except ValueError:
            print("  [오류] 숫자 또는 'N,M' 형식으로 입력하세요.")


# ── 전체 평가 ─────────────────────────────────────────────
def run_evaluation():
    print("[시작] QA 데이터셋 평가")
    print(f" - QA 데이터셋: {QA_DATASET_PATH}")
    print(f" - 하이브리드 모듈: {HYBRID_RAG_PATH}")

    module = load_hybrid_module()
    dataset = load_qa_dataset(QA_DATASET_PATH)

    collection_name = getattr(module, "COLLECTION_NAME", "")
    experiment_id = getattr(module, "EXPERIMENT_ID", "")

    # LLM-as-judge: 답변이 ground_truth 와 일치하는지 채점
    try:
        judge_llm = get_judge_llm()
        judge_model = os.getenv("EVAL_JUDGE_MODEL", JUDGE_MODEL)
        print(f" - Judge LLM: {judge_model} (Upstage, temp={JUDGE_TEMPERATURE})")
    except Exception as e:
        judge_llm = None
        print(f" - Judge LLM 사용 불가 ({type(e).__name__}: {e})")
        print(f"   → 정답 비교 없이 진행. 'success'는 실행 성공만 의미.")

    print(f" - 벡터 컬렉션: {collection_name}")
    print(f" - 실험 ID: {experiment_id}")
    print(f" - 총 질문 수: {len(dataset)}개")

    start_idx, end_idx = _prompt_eval_range(len(dataset))
    eval_subset = dataset[start_idx:end_idx]
    print(f" - 평가 범위: {start_idx + 1}~{end_idx}번 ({len(eval_subset)}개)")

    rows = []
    start_time = time.time()

    try:
        for local_idx, item in enumerate(eval_subset):
            idx = start_idx + local_idx
            question = extract_question(item)
            print(f"\n[{local_idx + 1}/{len(eval_subset)}] #{idx + 1} {question[:80]}")
            try:
                row = evaluate_one(module, item, idx, judge_llm)
            except Exception as e:
                row = {
                    "id": extract_id(item, idx),
                    "category": extract_category(item),
                    "question": question,
                    "ground_truth": extract_ground_truth(item),

                    "vector_success": False,
                    "vector_judge_verdict": "error",
                    "vector_judge_reason": "평가 핸들러 오류",
                    "vector_answer": "",
                    "vector_sources": "",
                    "vector_scored_sources": "",
                    "vector_context_preview": "",
                    "vector_error": f"Unhandled: {type(e).__name__}: {e}",

                    "hybrid_success": False,
                    "hybrid_judge_verdict": "error",
                    "hybrid_judge_reason": "평가 핸들러 오류",
                    "hybrid_answer": "",
                    "hybrid_sources": "",
                    "hybrid_scored_sources": "",
                    "hybrid_graph_count": 0,
                    "hybrid_graph_preview": "",
                    "hybrid_context_preview": "",
                    "hybrid_error": f"Unhandled: {type(e).__name__}: {e}",
                }

            row["no"] = f"{local_idx + 1:02d}"
            v_label = "정답" if row["vector_success"] else f"오답({row.get('vector_judge_verdict', '?')})"
            h_label = "정답" if row["hybrid_success"] else f"오답({row.get('hybrid_judge_verdict', '?')})"
            print(
                f"  - Vector: {v_label}"
                f" | Hybrid: {h_label}"
                f" | Graph relations: {row['hybrid_graph_count']}"
            )
            rows.append(row)
    except KeyboardInterrupt:
        print(f"\n\n[중단] Ctrl+C 감지 — {len(rows)}/{len(eval_subset)}개 완료. 결과 저장 중...")

    elapsed = round(time.time() - start_time, 2)

    if not rows:
        print("[종료] 저장할 결과 없음")
        return {}

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = os.path.join(RESULTS_DIR, f"evaluation_results_{timestamp}.xlsx")
    json_path = os.path.join(RESULTS_DIR, f"evaluation_results_{timestamp}.json")
    summary_path = os.path.join(RESULTS_DIR, f"evaluation_summary_{timestamp}.json")

    df = pd.DataFrame(rows)

    # 원본 json 먼저 저장 (xlsx 가 실패해도 raw 결과는 보존되도록 순서 변경)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f" - JSON 저장 완료: {json_path}")

    # xlsx 저장 (실패해도 JSON/summary 는 영향 없음)
    try:
        _save_clean_xlsx(xlsx_path, rows, collection_name, experiment_id, elapsed, timestamp)
    except Exception as e:
        print(f" - [xlsx 저장 실패] {type(e).__name__}: {str(e)[:200]}")
        print(f"   JSON 결과는 보존됨: {json_path}")

    def _count_verdict(field, target):
        if df.empty or field not in df.columns:
            return 0
        return int((df[field] == target).sum())

    summary = {
        "timestamp": timestamp,
        "total_questions": len(rows),
        "judge_used": judge_llm is not None,
        "vector_correct_count": int(df["vector_success"].sum()) if not df.empty else 0,
        "vector_incorrect_count": _count_verdict("vector_judge_verdict", "incorrect"),
        "vector_skipped_count": _count_verdict("vector_judge_verdict", "skipped"),
        "vector_error_count": _count_verdict("vector_judge_verdict", "error"),
        "hybrid_correct_count": int(df["hybrid_success"].sum()) if not df.empty else 0,
        "hybrid_incorrect_count": _count_verdict("hybrid_judge_verdict", "incorrect"),
        "hybrid_skipped_count": _count_verdict("hybrid_judge_verdict", "skipped"),
        "hybrid_error_count": _count_verdict("hybrid_judge_verdict", "error"),
        "avg_hybrid_graph_count": round(df["hybrid_graph_count"].mean(), 2) if not df.empty else 0,
        "collection_name": collection_name,
        "experiment_id": experiment_id,
        "elapsed_seconds": elapsed,
        "xlsx_path": xlsx_path,
        "json_path": json_path,
    }

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print("\n[완료] 평가 종료")
    n = summary["total_questions"] or 1
    print(f" - 총 질문 수: {summary['total_questions']}")
    if summary["judge_used"]:
        print(
            f" - Vector 정답: {summary['vector_correct_count']}/{n} "
            f"(오답 {summary['vector_incorrect_count']}, "
            f"스킵 {summary['vector_skipped_count']}, 에러 {summary['vector_error_count']})"
        )
        print(
            f" - Hybrid 정답: {summary['hybrid_correct_count']}/{n} "
            f"(오답 {summary['hybrid_incorrect_count']}, "
            f"스킵 {summary['hybrid_skipped_count']}, 에러 {summary['hybrid_error_count']})"
        )
    else:
        print(f" - Vector 실행 성공: {summary['vector_correct_count']}")
        print(f" - Hybrid 실행 성공: {summary['hybrid_correct_count']}")
        print(f" - (judge LLM 미사용 — 정답 비교 없음)")
    print(f" - 평균 Graph 관계 수: {summary['avg_hybrid_graph_count']}")
    print(f" - 소요 시간: {summary['elapsed_seconds']}초")
    print(f" - 결과 xlsx: {xlsx_path}")
    print(f" - 결과 json: {json_path}")

    return summary


if __name__ == "__main__":
    try:
        run_evaluation()
    except Exception as e:
        print("[치명 오류] 평가 중 예외 발생")
        print(f"{type(e).__name__}: {e}")
        print(traceback.format_exc())