"""
6개 모델 조합 Hybrid RAG 전체 평가 스크립트
==============================================
02_vector_db.py의 EXPERIMENTS와 동일한 6개 조합을 Hybrid RAG에도 적용하여
Vector Only / Hybrid 성능을 비교 평가합니다.

실험 조합:
  - openai_large  : text-embedding-3-large  + gpt-4o
  - openai_small  : text-embedding-3-small  + gpt-4o-mini
  - gemini_pro    : gemini-embedding-001    + gemini-1.5-pro
  - gemini_flash  : gemini-embedding-001    + gemini-1.5-flash
  - upstage_pro   : solar-embedding-passage + solar-pro
  - upstage_mini  : solar-embedding-passage + solar-mini

결과 저장:
  evaluation/results/evaluation_all_models_<timestamp>.xlsx
    - 시트1: 전체 요약   (6개 실험 비교)
    - 시트2~7: 실험별 상세 (openai_large, openai_small, ...)
    - 시트8: 전체 데이터  (110 * 6 = 660행)
  evaluation/results/evaluation_all_models_<timestamp>.json
  evaluation/results/evaluation_all_models_summary_<timestamp>.json
  evaluation/results/checkpoint/  (실험별 중간 저장)
"""

import os
import sys
import json
import time
import traceback
import importlib.util
from datetime import datetime

import pandas as pd
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 경로 설정 ─────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV_PATH = os.path.join(BASE_DIR, ".env")
EVAL_DIR = os.path.join(BASE_DIR, "evaluation")
RESULTS_DIR = os.path.join(EVAL_DIR, "results")
CHECKPOINT_DIR = os.path.join(RESULTS_DIR, "checkpoint")
QA_DATASET_PATH = os.path.join(EVAL_DIR, "qa_dataset.json")
CHROMA_DIR = os.path.join(BASE_DIR, "chroma_db")
HYBRID_RAG_PATH = os.path.join(BASE_DIR, "pipeline", "04_hybrid_rag.py")

os.makedirs(RESULTS_DIR, exist_ok=True)
os.makedirs(CHECKPOINT_DIR, exist_ok=True)
load_dotenv(ENV_PATH)

# ── 재시도 설정 ──────────────────────────────────────────
MAX_RETRIES = 3          # API 호출 최대 재시도 횟수
RETRY_DELAY = 5.0        # 초기 대기 시간 (초)
RETRY_BACKOFF = 2.0      # 대기 시간 배수 (지수 백오프)

# ── 6개 실험 조합 (02_vector_db.py와 동일) ───────────────
EXPERIMENTS = [
    {
        "id": "openai_large",
        "provider": "openai",
        "embedding_model": "text-embedding-3-large",
        "llm_model": "gpt-4o",
        "label": "OpenAI 대형 (text-embedding-3-large + gpt-4o)",
    },
    {
        "id": "openai_small",
        "provider": "openai",
        "embedding_model": "text-embedding-3-small",
        "llm_model": "gpt-4o-mini",
        "label": "OpenAI 소형 (text-embedding-3-small + gpt-4o-mini)",
    },
    {
        "id": "gemini_pro",
        "provider": "gemini",
        "embedding_model": "models/gemini-embedding-001",
        "llm_model": "gemini-1.5-pro",
        "label": "Gemini 대형 (gemini-embedding-001 + gemini-1.5-pro)",
    },
    {
        "id": "gemini_flash",
        "provider": "gemini",
        "embedding_model": "models/gemini-embedding-001",
        "llm_model": "gemini-1.5-flash",
        "label": "Gemini 소형 (gemini-embedding-001 + gemini-1.5-flash)",
    },
    {
        "id": "upstage_pro",
        "provider": "upstage",
        "embedding_model": "solar-embedding-1-large-passage",
        "llm_model": "solar-pro",
        "label": "Upstage 대형 (solar-embedding-passage + solar-pro)",
    },
    {
        "id": "upstage_mini",
        "provider": "upstage",
        "embedding_model": "solar-embedding-1-large-passage",
        "llm_model": "solar-mini",
        "label": "Upstage 소형 (solar-embedding-passage + solar-mini)",
    },
]


# ── [수정①] API 재시도 래퍼 ──────────────────────────────
def with_retry(fn, *args, label="API 호출", **kwargs):
    """지수 백오프 방식으로 최대 MAX_RETRIES번 재시도"""
    delay = RETRY_DELAY
    last_exc = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            last_exc = e
            err_str = str(e)
            # Rate Limit / 서버 오류 계열만 재시도
            is_retriable = any(
                kw in err_str.lower()
                for kw in ["rate limit", "429", "503", "timeout",
                           "resource exhausted", "quota", "connection"]
            )
            if not is_retriable or attempt == MAX_RETRIES:
                break
            print(f"    ⚠ {label} 실패 ({attempt}/{MAX_RETRIES}), {delay:.0f}초 후 재시도... [{type(e).__name__}]")
            time.sleep(delay)
            delay *= RETRY_BACKOFF
    raise last_exc


# ── [수정②] 실험별 클라이언트 캐시 (Chroma / LLM / Embedding) ──
_chroma_client_cache = {}
_embedding_fn_cache = {}


def get_chroma_client():
    """Chroma 클라이언트를 프로세스 내에서 1개만 유지"""
    if "client" not in _chroma_client_cache:
        import chromadb
        from chromadb.config import Settings
        _chroma_client_cache["client"] = chromadb.PersistentClient(
            path=CHROMA_DIR,
            settings=Settings(anonymized_telemetry=False),
        )
    return _chroma_client_cache["client"]


def get_embedding_function(experiment):
    """임베딩 함수를 실험 ID 단위로 캐싱 (매 호출마다 새로 생성 방지)"""
    exp_id = experiment["id"]
    if exp_id in _embedding_fn_cache:
        return _embedding_fn_cache[exp_id]

    provider = experiment["provider"]
    model = experiment["embedding_model"]

    if provider == "openai":
        from langchain_openai import OpenAIEmbeddings
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY가 .env에 없습니다.")
        fn = OpenAIEmbeddings(model=model, api_key=api_key)

    elif provider == "gemini":
        from langchain_google_genai import GoogleGenerativeAIEmbeddings
        api_key = os.getenv("GOOGLE_API_KEY")
        if not api_key:
            raise ValueError("GOOGLE_API_KEY가 .env에 없습니다.")
        fn = GoogleGenerativeAIEmbeddings(
            model=model,
            google_api_key=api_key,
            task_type="RETRIEVAL_DOCUMENT",
        )

    elif provider == "upstage":
        from langchain_upstage import UpstageEmbeddings
        api_key = os.getenv("UPSTAGE_API_KEY")
        if not api_key:
            raise ValueError("UPSTAGE_API_KEY가 .env에 없습니다.")
        fn = UpstageEmbeddings(model=model, api_key=api_key)

    else:
        raise ValueError(f"지원하지 않는 provider: {provider}")

    _embedding_fn_cache[exp_id] = fn
    return fn


# ── LLM 답변 생성 ─────────────────────────────────────────
ANSWER_PROMPT_TEMPLATE = """당신은 경북대학교 컴퓨터학부 학생들을 위한 AI 챗봇입니다.
아래 검색된 문서 내용과 그래프 관계 정보를 바탕으로 질문에 답변하세요.
반드시 검색 결과에 근거해서만 답변하세요.
검색 결과에 없는 내용은 추측하지 말고 "해당 정보를 찾을 수 없습니다."라고 답변하세요.

[검색 결과]
{context}

[질문]
{query}

[답변]
"""


def _call_llm(provider, model, prompt):
    """LLM 실제 호출 (with_retry에서 감싸서 사용)"""
    if provider == "openai":
        from openai import OpenAI
        client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        )
        return response.choices[0].message.content.strip()

    elif provider == "gemini":
        from langchain_google_genai import ChatGoogleGenerativeAI
        llm = ChatGoogleGenerativeAI(
            model=model,
            google_api_key=os.getenv("GOOGLE_API_KEY"),
            temperature=0.2,
        )
        return llm.invoke(prompt).content.strip()

    elif provider == "upstage":
        from langchain_upstage import ChatUpstage
        llm = ChatUpstage(
            model=model,
            api_key=os.getenv("UPSTAGE_API_KEY"),
            temperature=0.2,
        )
        return llm.invoke(prompt).content.strip()

    raise ValueError(f"지원하지 않는 provider: {provider}")


def generate_answer(query, context, experiment):
    """재시도 포함 LLM 답변 생성"""
    prompt = ANSWER_PROMPT_TEMPLATE.format(context=context, query=query)
    return with_retry(
        _call_llm,
        experiment["provider"],
        experiment["llm_model"],
        prompt,
        label=f"LLM({experiment['llm_model']})",
    )


# ── Vector 검색 ───────────────────────────────────────────
def _do_embed_query(embedding_fn, query):
    return embedding_fn.embed_query(query)


def vector_search(query, experiment, n_results=3):
    """캐싱된 Chroma 클라이언트로 유사 문서 검색 (재시도 포함)"""
    exp_id = experiment["id"]
    collection_name = f"knu_cse_{exp_id}"

    client = get_chroma_client()
    existing = client.list_collections()
    if collection_name not in existing:
        raise ValueError(
            f"컬렉션 없음: {collection_name}\n"
            f"02_vector_db.py를 먼저 실행해서 벡터 DB를 구축하세요.\n"
            f"현재 존재하는 컬렉션: {existing}"
        )

    collection = client.get_collection(collection_name)
    embedding_fn = get_embedding_function(experiment)

    query_embedding = with_retry(
        _do_embed_query,
        embedding_fn,
        query,
        label=f"임베딩({experiment['embedding_model']})",
    )

    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=n_results,
    )

    docs = []
    for doc, meta, dist in zip(
        results.get("documents", [[]])[0],
        results.get("metadatas", [[]])[0],
        results.get("distances", [[]])[0],
    ):
        score = round(1 - dist, 3) if dist is not None else None
        docs.append({
            "content": doc,
            "source": meta.get("file_name", "") if meta else "",
            "score": score,
        })

    return docs


# ── 결과 병합 ─────────────────────────────────────────────
def merge_results(vector_docs, graph_relations):
    context_parts = []

    if vector_docs:
        vector_text = "=== 문서 검색 결과 ===\n"
        for i, doc in enumerate(vector_docs, start=1):
            vector_text += f"\n[{i}] 출처: {doc['source']} (유사도: {doc['score']})\n"
            vector_text += f"{doc['content']}\n"
        context_parts.append(vector_text.strip())

    if graph_relations:
        graph_text = "=== 관계 그래프 검색 결과 ===\n"
        for rel in graph_relations:
            if rel.get("from") and rel.get("to"):
                graph_text += f"- {rel['from']} --[{rel['relation']}]--> {rel['to']}\n"
        context_parts.append(graph_text.strip())

    return "\n\n".join(context_parts).strip()


# ── 04_hybrid_rag.py 동적 로드 ────────────────────────────
def load_hybrid_module():
    spec = importlib.util.spec_from_file_location("hybrid_rag_module", HYBRID_RAG_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


# ── QA 데이터셋 유틸 ──────────────────────────────────────
def load_qa_dataset():
    if not os.path.exists(QA_DATASET_PATH):
        raise FileNotFoundError(f"QA 데이터셋이 없습니다: {QA_DATASET_PATH}")
    with open(QA_DATASET_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("qa_dataset.json은 list 형태여야 합니다.")
    return data


def extract_question(item):
    for key in ["question", "query", "질문"]:
        v = item.get(key)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def extract_ground_truth(item):
    for key in ["ground_truth", "answer", "reference_answer", "expected_answer", "정답"]:
        v = item.get(key)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def extract_id(item, index):
    for key in ["id", "qid", "question_id", "번호"]:
        v = item.get(key)
        if v is not None and str(v).strip():
            return str(v).strip()
    return str(index + 1)


def extract_category(item):
    for key in ["category", "type", "persona", "source", "분류"]:
        v = item.get(key)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def truncate_text(text, limit=1500):
    if not text:
        return ""
    text = str(text)
    return text[:limit] + "...[truncated]" if len(text) > limit else text


def safe_join(items, sep=" | "):
    return sep.join([str(x) for x in items if x is not None and str(x).strip()])


def summarize_vector_docs(vector_docs):
    sources, scored = [], []
    for doc in vector_docs or []:
        src = doc.get("source", "")
        score = doc.get("score", "")
        if src:
            sources.append(src)
            scored.append(f"{src} ({score})")
    return {
        "sources": safe_join(dict.fromkeys(sources).keys()),
        "scored_sources": safe_join(scored),
    }


def summarize_graph_relations(graph_relations, max_preview=10):
    preview = []
    for rel in (graph_relations or [])[:max_preview]:
        f, r, t = rel.get("from", ""), rel.get("relation", ""), rel.get("to", "")
        if f and r and t:
            preview.append(f"{f} --[{r}]--> {t}")
    return {
        "count": len(graph_relations or []),
        "preview": "\n".join(preview),
    }


# ── 질문 1개 평가 ─────────────────────────────────────────
def evaluate_one(item, index, experiment, hybrid_module):
    question = extract_question(item)
    qid = extract_id(item, index)
    category = extract_category(item)
    ground_truth = extract_ground_truth(item)

    # [수정⑤] vector_success를 처음부터 문자열로 초기화
    row = {
        "id": qid,
        "category": category,
        "question": question,
        "ground_truth": ground_truth,
        "experiment_id": experiment["id"],
        "experiment_label": experiment["label"],
        "provider": experiment["provider"],
        "embedding_model": experiment["embedding_model"],
        "llm_model": experiment["llm_model"],

        "vector_success": "실패",
        "vector_answer": "",
        "vector_sources": "",
        "vector_scored_sources": "",
        "vector_context_preview": "",
        "vector_error": "",

        "hybrid_success": "실패",
        "hybrid_answer": "",
        "hybrid_sources": "",
        "hybrid_scored_sources": "",
        "hybrid_graph_count": 0,
        "hybrid_graph_preview": "",
        "hybrid_context_preview": "",
        "hybrid_error": "",
    }

    if not question:
        row["vector_error"] = "질문 없음"
        row["hybrid_error"] = "질문 없음"
        return row

    # ── [수정③] Vector 검색 1번만 수행 후 두 블록에서 공유 ──
    v_docs = None
    v_search_error = None
    try:
        v_docs = vector_search(question, experiment, n_results=3)
    except Exception as e:
        v_search_error = f"{type(e).__name__}: {e}"

    # ── Vector Only ───────────────────────────────────────
    try:
        if v_docs is None:
            raise RuntimeError(v_search_error or "Vector 검색 실패")
        context = merge_results(v_docs, [])
        answer = generate_answer(question, context, experiment)
        v_sum = summarize_vector_docs(v_docs)

        row["vector_success"] = "성공"
        row["vector_answer"] = answer
        row["vector_sources"] = v_sum["sources"]
        row["vector_scored_sources"] = v_sum["scored_sources"]
        row["vector_context_preview"] = truncate_text(context)
    except Exception as e:
        row["vector_error"] = f"{type(e).__name__}: {e}"

    # ── Hybrid (Vector + Graph) ───────────────────────────
    try:
        if v_docs is None:
            raise RuntimeError(v_search_error or "Vector 검색 실패")
        graph_rels = with_retry(
            hybrid_module.graph_search,
            question,
            label="Graph 검색(Neo4j)",
        )
        context = merge_results(v_docs, graph_rels)
        answer = generate_answer(question, context, experiment)
        v_sum = summarize_vector_docs(v_docs)
        g_sum = summarize_graph_relations(graph_rels)

        row["hybrid_success"] = "성공"
        row["hybrid_answer"] = answer
        row["hybrid_sources"] = v_sum["sources"]
        row["hybrid_scored_sources"] = v_sum["scored_sources"]
        row["hybrid_graph_count"] = g_sum["count"]
        row["hybrid_graph_preview"] = g_sum["preview"]
        row["hybrid_context_preview"] = truncate_text(context)
    except Exception as e:
        row["hybrid_error"] = f"{type(e).__name__}: {e}"

    return row


# ── 실험 1개 전체 평가 ────────────────────────────────────
def run_experiment(experiment, dataset, hybrid_module, timestamp):
    label = experiment["label"]
    exp_id = experiment["id"]
    rows = []

    print(f"\n{'=' * 65}")
    print(f"[실험] {label}")
    print(f"{'=' * 65}")

    for idx, item in enumerate(dataset):
        question = extract_question(item)
        print(f"  [{idx + 1:>3}/{len(dataset)}] {question[:65]}")

        try:
            row = evaluate_one(item, idx, experiment, hybrid_module)
        except Exception as e:
            row = {
                "id": extract_id(item, idx),
                "category": extract_category(item),
                "question": question,
                "ground_truth": extract_ground_truth(item),
                "experiment_id": exp_id,
                "experiment_label": experiment["label"],
                "provider": experiment["provider"],
                "embedding_model": experiment["embedding_model"],
                "llm_model": experiment["llm_model"],
                "vector_success": "실패", "vector_answer": "",
                "vector_sources": "", "vector_scored_sources": "",
                "vector_context_preview": "",
                "vector_error": f"Unhandled: {type(e).__name__}: {e}",
                "hybrid_success": "실패", "hybrid_answer": "",
                "hybrid_sources": "", "hybrid_scored_sources": "",
                "hybrid_graph_count": 0, "hybrid_graph_preview": "",
                "hybrid_context_preview": "",
                "hybrid_error": f"Unhandled: {type(e).__name__}: {e}",
            }

        v_ok = row["vector_success"] == "성공"
        h_ok = row["hybrid_success"] == "성공"
        print(
            f"         Vector: {'✓' if v_ok else '✗'}"
            f" | Hybrid: {'✓' if h_ok else '✗'}"
            f" | Graph: {row['hybrid_graph_count']}개"
        )
        rows.append(row)

    # ── [수정④] 실험 완료 시 즉시 체크포인트 저장 ────────
    ckpt_path = os.path.join(CHECKPOINT_DIR, f"{exp_id}_{timestamp}.json")
    try:
        with open(ckpt_path, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        print(f"  [체크포인트 저장] {ckpt_path}")
    except Exception as e:
        print(f"  [체크포인트 저장 실패] {e}")

    return rows


# ── [수정②] 실험별 성공률 계산 (문자열 기준) ─────────────
def compute_summary(experiment, rows, elapsed):
    total = len(rows)
    v_success = sum(1 for r in rows if r.get("vector_success") == "성공")
    h_success = sum(1 for r in rows if r.get("hybrid_success") == "성공")
    avg_graph = (
        sum(r.get("hybrid_graph_count", 0) for r in rows) / total
        if total > 0 else 0.0
    )
    return {
        "experiment_id": experiment["id"],
        "experiment_label": experiment["label"],
        "provider": experiment["provider"],
        "embedding_model": experiment["embedding_model"],
        "llm_model": experiment["llm_model"],
        "total_questions": total,
        "vector_success_count": v_success,
        "hybrid_success_count": h_success,
        "vector_success_rate(%)": round(v_success / total * 100, 1) if total > 0 else 0.0,
        "hybrid_success_rate(%)": round(h_success / total * 100, 1) if total > 0 else 0.0,
        "avg_graph_count": round(avg_graph, 2),
        "elapsed_seconds": elapsed,
    }


# ── 모델 짧은 레이블 ─────────────────────────────────────
SHORT_LABEL = {
    "openai_large": "GPT-4o",
    "openai_small": "GPT-4o-mini",
    "gemini_pro":   "Gemini-Pro",
    "gemini_flash": "Gemini-Flash",
    "upstage_pro":  "Solar-Pro",
    "upstage_mini": "Solar-Mini",
}


# ── xlsx 저장 (새 구조) ───────────────────────────────────
def _save_xlsx(all_rows, experiment_summaries, xlsx_path):
    """
    시트 구성:
      ① 요약          : 6개 모델 성능 비교 (7열)
      ② Vector 비교   : 질문별 6모델 답변 나란히 (10열)
      ③ Hybrid 비교   : 질문별 6모델 답변+Graph수 나란히 (16열)
      ④~⑨ 모델별 상세: 핵심 열만 (11열, 기존 23열에서 축소)
    """
    # ── 색상 팔레트 ──────────────────────────────────────
    C = {
        "h_blue":    PatternFill("solid", fgColor="1F497D"),  # 상세 헤더
        "h_teal":    PatternFill("solid", fgColor="17375E"),  # 비교 헤더
        "h_green":   PatternFill("solid", fgColor="375623"),  # 요약 헤더
        "success":   PatternFill("solid", fgColor="E2EFDA"),  # 성공 셀
        "fail":      PatternFill("solid", fgColor="FFDDC1"),  # 실패 셀
        "q_bg":      PatternFill("solid", fgColor="F2F2F2"),  # 질문 배경
        "none":      PatternFill("solid", fgColor="FFFFFF"),
    }
    F = {
        "header": Font(bold=True, color="FFFFFF", name="Arial", size=9),
        "bold":   Font(bold=True, name="Arial", size=9),
        "normal": Font(name="Arial", size=9),
        "small":  Font(name="Arial", size=8, color="666666"),
    }
    AL_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_TL = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    AL_TC = Alignment(horizontal="center", vertical="top")

    # ── 데이터 그룹화 ─────────────────────────────────────
    from collections import defaultdict
    by_exp = defaultdict(list)
    for r in all_rows:
        by_exp[r["experiment_id"]].append(r)

    # 질문 순서 고정 (openai_small 기준, 없으면 첫 실험)
    base_exp = "openai_small" if "openai_small" in by_exp else EXPERIMENTS[0]["id"]
    base_rows = by_exp[base_exp]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # ── ① 요약 시트 ──────────────────────────────────────
    ws = wb.create_sheet("① 요약")
    headers = ["모델", "임베딩", "LLM", "Vector 성공률", "Hybrid 성공률", "평균 Graph 수", "소요(분)"]
    ws.append(headers)
    for s in experiment_summaries:
        ws.append([
            SHORT_LABEL.get(s["experiment_id"], s["experiment_id"]),
            s["embedding_model"],
            s["llm_model"],
            f"{s['vector_success_rate(%)']}%",
            f"{s['hybrid_success_rate(%)']}%",
            s["avg_graph_count"],
            round(s["elapsed_seconds"] / 60, 1),
        ])
    _fmt_sheet(ws, C["h_green"], F, AL_C, AL_TL,
               col_widths=[14, 32, 18, 14, 14, 14, 10],
               freeze="A2", success_col=None)

    # ── ② Vector 비교 시트 ───────────────────────────────
    ws = wb.create_sheet("② Vector 비교")
    model_ids = [e["id"] for e in EXPERIMENTS]
    v_headers = ["번호", "분류", "질문", "정답"] + [SHORT_LABEL.get(m, m) for m in model_ids]
    ws.append(v_headers)

    # 답변 열 인덱스 (1-based, 5번째부터)
    ans_col_start = 5
    for i, base_row in enumerate(base_rows):
        qid      = base_row["id"]
        category = base_row["category"]
        question = base_row["question"]
        gt       = base_row["ground_truth"]
        answers  = []
        successes = []
        for exp_id in model_ids:
            exp_rows = by_exp[exp_id]
            matched  = next((r for r in exp_rows if r["id"] == qid), None)
            if matched:
                answers.append(matched.get("vector_answer", ""))
                successes.append(matched.get("vector_success", "실패"))
            else:
                answers.append("데이터 없음")
                successes.append("실패")
        ws.append([qid, category, question, gt] + answers)
        # 성공/실패 배경색
        row_idx = i + 2
        for col_offset, success in enumerate(successes):
            cell = ws.cell(row=row_idx, column=ans_col_start + col_offset)
            cell.fill = C["success"] if success == "성공" else C["fail"]
            cell.alignment = AL_TL
            cell.font = F["normal"]

    _fmt_sheet(ws, C["h_teal"], F, AL_C, AL_TL,
               col_widths=[6, 10, 40, 30] + [35] * 6,
               freeze="E2", success_col=None, q_col=3)

    # ── ③ Hybrid 비교 시트 ───────────────────────────────
    ws = wb.create_sheet("③ Hybrid 비교")
    h_headers = ["번호", "분류", "질문", "정답"]
    for m in model_ids:
        label = SHORT_LABEL.get(m, m)
        h_headers += [f"{label} 답변", f"{label} Graph수"]
    ws.append(h_headers)

    for i, base_row in enumerate(base_rows):
        qid      = base_row["id"]
        category = base_row["category"]
        question = base_row["question"]
        gt       = base_row["ground_truth"]
        cells_data = []
        successes  = []
        for exp_id in model_ids:
            exp_rows = by_exp[exp_id]
            matched  = next((r for r in exp_rows if r["id"] == qid), None)
            if matched:
                cells_data += [matched.get("hybrid_answer", ""), matched.get("hybrid_graph_count", 0)]
                successes.append(matched.get("hybrid_success", "실패"))
            else:
                cells_data += ["데이터 없음", 0]
                successes.append("실패")
        ws.append([qid, category, question, gt] + cells_data)
        row_idx = i + 2
        for col_offset, success in enumerate(successes):
            ans_cell   = ws.cell(row=row_idx, column=5 + col_offset * 2)
            graph_cell = ws.cell(row=row_idx, column=6 + col_offset * 2)
            ans_cell.fill   = C["success"] if success == "성공" else C["fail"]
            graph_cell.fill = C["success"] if success == "성공" else C["fail"]
            ans_cell.alignment   = AL_TL
            graph_cell.alignment = AL_TC
            ans_cell.font   = F["normal"]
            graph_cell.font = F["normal"]

    col_widths = [6, 10, 40, 30] + [35, 10] * 6
    _fmt_sheet(ws, C["h_teal"], F, AL_C, AL_TL,
               col_widths=col_widths, freeze="E2", success_col=None, q_col=3)

    # ── ④~⑨ 모델별 상세 시트 (11열) ─────────────────────
    DETAIL_COLS = [
        ("번호",        "id"),
        ("분류",        "category"),
        ("질문",        "question"),
        ("정답",        "ground_truth"),
        ("Vector 성공", "vector_success"),
        ("Vector 답변", "vector_answer"),
        ("Vector 오류", "vector_error"),
        ("Hybrid 성공", "hybrid_success"),
        ("Hybrid 답변", "hybrid_answer"),
        ("Graph 수",    "hybrid_graph_count"),
        ("Hybrid 오류", "hybrid_error"),
    ]
    detail_headers = [h for h, _ in DETAIL_COLS]
    detail_keys    = [k for _, k in DETAIL_COLS]

    for exp in EXPERIMENTS:
        short = SHORT_LABEL.get(exp["id"], exp["id"])
        ws = wb.create_sheet(f"④ {short}"[:31] if exp == EXPERIMENTS[0]
                             else short[:31])
        # 시트 번호 붙이기
        sheet_num = EXPERIMENTS.index(exp) + 4
        ws.title = f"{sheet_num:01d}. {short}"[:31]

        ws.append(detail_headers)
        for r in by_exp[exp["id"]]:
            ws.append([r.get(k, "") for k in detail_keys])

        _fmt_sheet(ws, C["h_blue"], F, AL_C, AL_TL,
                   col_widths=[6, 10, 40, 30, 10, 40, 25, 10, 40, 8, 25],
                   freeze="E2",
                   success_col=[5, 8],  # Vector성공, Hybrid성공 열 (1-based)
                   q_col=3)

    wb.save(xlsx_path)


def _fmt_sheet(ws, header_fill, F, AL_C, AL_TL,
               col_widths=None, freeze="A2", success_col=None, q_col=None):
    """공통 시트 서식 적용"""
    SUCCESS_FILL = PatternFill("solid", fgColor="E2EFDA")
    FAIL_FILL    = PatternFill("solid", fgColor="FFDDC1")
    Q_FILL       = PatternFill("solid", fgColor="F5F5F5")

    # 헤더
    for cell in ws[1]:
        cell.font      = F["header"]
        cell.fill      = header_fill
        cell.alignment = AL_C
    ws.row_dimensions[1].height = 32

    # 데이터
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 15
        for col_idx, cell in enumerate(row, start=1):
            cell.font      = F["normal"]
            cell.alignment = AL_TL
            # 성공/실패 셀 색상 (지정된 열만)
            if success_col and col_idx in success_col:
                if cell.value == "성공":
                    cell.fill      = SUCCESS_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif cell.value == "실패":
                    cell.fill      = FAIL_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="top")
            # 질문 열 배경
            if q_col and col_idx == q_col:
                cell.fill = Q_FILL

    # 열 너비
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col_idx, col in enumerate(ws.iter_cols(), start=1):
            max_len = max((min(len(str(c.value or "")), 60) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(8, max_len + 2)

    if freeze:
        ws.freeze_panes = freeze


# ── 전체 평가 실행 ────────────────────────────────────────
def run_all_evaluation():
    print("=" * 65)
    print("[시작] 6개 모델 조합 Hybrid RAG 전체 평가")
    print("=" * 65)
    print(f"  QA 데이터셋 : {QA_DATASET_PATH}")
    print(f"  실험 수     : {len(EXPERIMENTS)}개")
    print(f"  API 재시도  : 최대 {MAX_RETRIES}회 (대기 {RETRY_DELAY}초 ~ 지수 백오프)")

    hybrid_module = load_hybrid_module()
    dataset = load_qa_dataset()
    print(f"  총 질문 수  : {len(dataset)}개")
    print(f"  총 평가 수  : {len(EXPERIMENTS) * len(dataset)}개 (실험 × 질문)")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    all_rows = []
    experiment_summaries = []
    total_start = time.time()

    for experiment in EXPERIMENTS:
        exp_start = time.time()
        rows = run_experiment(experiment, dataset, hybrid_module, timestamp)
        exp_elapsed = round(time.time() - exp_start, 2)

        all_rows.extend(rows)
        summary = compute_summary(experiment, rows, exp_elapsed)
        experiment_summaries.append(summary)

        print(f"\n  ▶ {experiment['label']}")
        print(f"    Vector : {summary['vector_success_count']}/{summary['total_questions']} ({summary['vector_success_rate(%)']}%)")
        print(f"    Hybrid : {summary['hybrid_success_count']}/{summary['total_questions']} ({summary['hybrid_success_rate(%)']}%)")
        print(f"    평균 Graph : {summary['avg_graph_count']}개 | 소요 : {exp_elapsed}초")

    total_elapsed = round(time.time() - total_start, 2)

    # ── 결과 저장 ─────────────────────────────────────────
    xlsx_path = os.path.join(RESULTS_DIR, f"evaluation_all_models_{timestamp}.xlsx")
    json_path = os.path.join(RESULTS_DIR, f"evaluation_all_models_{timestamp}.json")
    summary_path = os.path.join(RESULTS_DIR, f"evaluation_all_models_summary_{timestamp}.json")

    # xlsx 저장 (새 구조)
    _save_xlsx(all_rows, experiment_summaries, xlsx_path)

    # JSON 저장
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)

    final_summary = {
        "timestamp": timestamp,
        "total_experiments": len(EXPERIMENTS),
        "questions_per_experiment": len(dataset),
        "total_evaluations": len(all_rows),
        "total_elapsed_seconds": total_elapsed,
        "experiments": experiment_summaries,
        "xlsx_path": xlsx_path,
        "json_path": json_path,
    }

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(final_summary, f, ensure_ascii=False, indent=2)

    # ── 최종 결과 출력 ────────────────────────────────────
    print(f"\n{'=' * 65}")
    print("[완료] 전체 평가 종료")
    print(f"  총 소요 시간 : {total_elapsed}초 ({round(total_elapsed / 60, 1)}분)")
    print(f"  xlsx 결과   : {xlsx_path}")
    print(f"  json 결과   : {json_path}")
    print(f"\n[결과 요약]")
    print(f"  {'실험':45s} {'Vector':>8} {'Hybrid':>8} {'평균Graph':>10}")
    print(f"  {'-'*45} {'-'*8} {'-'*8} {'-'*10}")
    for s in experiment_summaries:
        label_short = s["experiment_label"][:45]
        print(
            f"  {label_short:45s}"
            f" {s['vector_success_rate(%)']:>7}%"
            f" {s['hybrid_success_rate(%)']:>7}%"
            f" {s['avg_graph_count']:>9}개"
        )

    return final_summary


if __name__ == "__main__":
    try:
        run_all_evaluation()
    except Exception as e:
        print(f"\n[치명 오류] {type(e).__name__}: {e}")
        print(traceback.format_exc())
        sys.exit(1)
